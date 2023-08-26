import os

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from .Tourbus import Tourbus, Tourist
from colour import Color
import string
import json



def getExcelCol(num: int):
    alphabet = string.ascii_uppercase
    numLetters = 26
    col = ""
    while num > 0:
        remainder = (num - 1) % numLetters
        num = (num - remainder) // numLetters
        col = alphabet[remainder] + col
    return col


def colorPicker(num, maxNum):
    normalized = (float(num) / maxNum) * 6.0
    if normalized < 1:
        r = 1
        b = 0
        g = normalized
    elif normalized < 2:
        r = 1 - (normalized - 1)
        b = 0
        g = 1
    elif normalized < 3:
        r = 0
        b = normalized - 2
        g = 1
    elif normalized < 4:
        r = 0
        b = 1
        g = 1 - (normalized - 3)
    elif normalized < 5:
        r = normalized - 4
        b = 1
        g = 0
    else:
        r = 1
        g = 0
        b = 1 - (normalized - 5)
    c = Color(rgb=(r, g, b))
    c.set_luminance(0.75)
    return f"00{c.get_hex_l()[1:]}"



def makeExcelFile(data: json):
    try:
        numDays = int(data["numDays"])
        tourists = []

        if "groupedTourists" in data:
            for group in data["groupedTourists"]:
                groupID = group["groupID"]
                tourists.extend([Tourist(i, groupID=groupID) for i in group["tourists"]])
        if "tourists" in data:
            tourists.extend([Tourist(i) for i in data["tourists"]])
    except KeyError:
        errorStr = ""
        errorStr += "Json file not set up correctly\n"
        errorStr += "Example of correct format:\n"
        with open(os.path.join(dirPath, 'sampleTourists.json')) as file:
            errorStr += json.dumps(json.load(file))
        raise KeyError(errorStr)

    tourbus = Tourbus(tourists, numDays)
    tourbus.fillSeatsForTrip()
    tourbus.addOneToAllSeatPositions()
    tourists = tourbus.getTourists()
    tourists = sorted(tourists, key=lambda h: (h.seatPositions[0]))
    fileName = "tourbus.xlsx"
    cwd = os.path.abspath(os.path.dirname(__file__))
    # absPathToFile = os.path.join(cwd, path)
    pathToFile = os.path.join(cwd, "../app/src", fileName)
    # pathToFile = f"{cwd}../app/{fileName}"

    workbook = Workbook()
    summaryPage = workbook.active
    touristColLength = max([len(i) for i in [i.name for i in tourists] + ["tourists"]])
    summaryPage.column_dimensions["A"].width = touristColLength + 1

    summaryPage["A1"] = "Tourists"
    for daySheet in range(numDays):
        summaryPage[f"{getExcelCol(daySheet + 2)}1"] = f"Day {daySheet + 1}"
    summaryPage[f"{getExcelCol(numDays + 2)}1"] = f"Seat Score:"

    for row in range(len(tourists)):
        tourist = tourists[row]
        excelRow = row + 2
        cell = summaryPage[f"A{excelRow}"]
        cell.value = tourist.name
        color = colorPicker(row, len(tourists))
        cell.fill = PatternFill(fgColor=color, fill_type="solid")
        for col in range(len(tourist.seatPositions)):
            seat = tourist.seatPositions[col]
            cell = summaryPage[f"{getExcelCol(col + 2)}{excelRow}"]
            cell.value = seat

        summaryPage[f"{getExcelCol(len(tourist.seatPositions) + 2)}{excelRow}"] = tourist.totalSeatScore

    BUS_ROW_OFFSET = 4
    BUS_COL_OFFSET = 3
    for i in range(numDays):
        dayNum = i + 1
        daySheet = workbook.create_sheet(f"Day {dayNum}")
        daySheet["A1"] = "Tourists"
        daySheet.column_dimensions["A"].width = touristColLength + 1
        daySheet["B1"] = "Seat Position"
        daySheet.column_dimensions["B"].width = len("Seat Position")

        daySheet[f"{getExcelCol(BUS_COL_OFFSET + 1)}1"] = "Bus Layout for the Day:"

        daySheet.column_dimensions[f"{getExcelCol(BUS_COL_OFFSET + 1)}"].width = touristColLength + 1
        daySheet.column_dimensions[f"{getExcelCol(BUS_COL_OFFSET + 2)}"].width = touristColLength + 1
        for row in range(len(tourists)):
            tourist = tourists[row]
            excelRow = row + 2
            cell = daySheet[f"A{excelRow}"]
            cell.value = tourist.name
            color = colorPicker(row, len(tourists))
            cell.fill = PatternFill(fgColor=color, fill_type="solid")
            daySheet[f"B{excelRow}"] = tourist.seatPositions[i]

            row, col = tourbus.getRowAndCol(tourist.seatPositions[i] - 1)
            busSeatCell = daySheet[f"{getExcelCol(col + BUS_ROW_OFFSET)}{row + BUS_COL_OFFSET}"]
            busSeatCell.fill = PatternFill(fgColor=color, fill_type="solid")
            busSeatCell.value = tourist.name
    print(pathToFile)
    workbook.save(pathToFile)
    # print(os.path.join(os.getcwd(), fileName))


if __name__ == "__main__":
    dirPath = os.path.dirname(os.path.abspath(__file__))
    try:
        with open(os.path.join(dirPath, 'tourists.json')) as file:
            data = json.load(file)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found, need a file named tourists.json in {dirPath}")
    except json.JSONDecodeError:
        raise json.JSONDecodeError("Error decoding JSON file")

    makeExcelFile(data)
